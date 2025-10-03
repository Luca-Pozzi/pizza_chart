import os
import pandas as pd

if __name__ == '__main__':
    # Define paths to input CSV and output Excel file
    csv_in_path = 'input/data/upload.csv'
    csv_out_path = 'data/data.csv'
    excel_out_path = csv_out_path.replace('.csv', '.xlsx')
    # Load data
    df_in = pd.read_csv(csv_in_path)
    cols_in = [col.strip() for col in df_in.columns]
    # Validate structure
    cols_required = {'Date', 'Pizza', '#', 'Pizzeria'}
    if not cols_required.issubset(cols_in):
        raise Exception(f'Missing columns. Required: {cols_required}. Found: {cols_in}')
    else:
        # Strip leading/trailing spaces from all string columns
        df_in.columns = cols_in  # update DataFrame columns
        for col in df_in.select_dtypes(include='object').columns:
            df_in[col] = df_in[col].str.strip()
        # Standardize Date column to dd/mm/yyyy with leading zeros
        df_in['Date'] = pd.to_datetime(df_in['Date'], 
                                       dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y'
                                       )
    if os.path.exists(csv_out_path): # data in CSV
        # Append to CSV
        if os.path.getsize(csv_out_path) > 0:
            existing_df = pd.read_csv(csv_out_path)
            df_out = pd.concat([existing_df, df_in], ignore_index=True)
            df_out = df_out.drop_duplicates()
        else:
            df_out = df_in
        df_out.to_csv(csv_out_path, index=False)
        print(f"Appended data to {csv_out_path}")
    elif os.path.exists(excel_out_path): # data in Excel
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        # Open Excel workbook and sheet
        wb = load_workbook(excel_out_path)
        wb.iso_dates = True
        ws = wb['Orders']
        # Find max row manually as `ws.max_row` may be incorrect if there are empty rows with some formatting.
        for cell in ws['A']:
            if cell.value is None:
                max_row = cell.row - 1
                break
        # Append rows
        fmt_dict = {
            'Date': {'number_format': 'DD/MM/YYYY', 'alignment': 'right'},
            'Pizza': {'alignment': 'left'},
            '#': {'number_format': '0', 'alignment': 'center'},
            'Pizzeria': {'alignment': 'left'}
        }
        next_row = max_row + 1 # first empty row
        for _, row in df_in.iterrows():
            for col_idx, value in enumerate(row.tolist(), 1):
                # Set the new cell value
                cell = ws.cell(row=next_row, column=col_idx)
                cell.value=value.strip() if isinstance(value, str) else value
                # Apply formatting based on column
                col_fmt_dict = fmt_dict.get(cols_in[col_idx-1], {})
                number_format = col_fmt_dict.get('number_format', None)
                if number_format:
                    cell.number_format=number_format
                cell.alignment = Alignment(horizontal=col_fmt_dict.get('alignment',
                                                                        'left'))
            next_row += 1
        # Extend the table range
        ws.tables['Orders'].ref = f"A1:D{next_row-1}"
        # Extend the conditional formatting range
        # TODO. Check if there is already a conditional formatting rule and update it. Looking from Excel, the newly appended lines are already included in the existing rules, but are -for some reason- not formatted.
        # Save workbook
        wb.save(excel_out_path)
    else:
        raise Exception(f"No data file found at {csv_out_path} or {excel_out_path}")